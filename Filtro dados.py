import sys, os
from PyQt5.QtWidgets import QApplication,QLabel,QComboBox, QVBoxLayout, QWidget, QPushButton, QFileDialog, QTableWidget, QTableWidgetItem,QLineEdit,QHBoxLayout
import pandas as pd
from openpyxl import Workbook

class MainWindow(QWidget):
    def __init__(self):
            super().__init__()

            self.label = QLabel("Selecione o bairro:")

            self.combobox = QComboBox()
            self.combobox.addItem("Todos os bairros")  # Add an option for all neighborhoods
            
            self.combobox.activated.connect(self.filtrar)

            
            

            self.line_edit = QLineEdit()

            # Botão de salvar
            self.save_button = QPushButton("Salvar")
            self.save_button.clicked.connect(self.salvar_dados)
            # Connect the save button to your desired saving functionality (placeholder)
            #self.save_button.clicked.connect(self.save_data)

            # Table
            self.table = QTableWidget()

            layout = QVBoxLayout()
            layout.addWidget(self.label)

            hbox = QHBoxLayout()
            hbox.addWidget(self.combobox)
            hbox.addWidget(self.line_edit)
            hbox.addWidget(self.save_button)
            layout.addLayout(hbox)

            self.button = QPushButton("Abrir Arquivo")
            self.button.clicked.connect(self.abrir_arquivo)
            layout.addWidget(self.button)

            layout.addWidget(self.table)
            self.setLayout(layout)

            self.show()


    def abrir_arquivo(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        self.combobox.clear()

        if hasattr(QFileDialog, 'filters'):  # Check for 'filters' attribute (PyQt5)
            filters = [("CSV Files (*.csv)", "*.csv"), ("Excel Files (*.xlsx *.xlsm)", "*.xlsx *.xlsm")]
            file_name, _ = QFileDialog.getOpenFileName(self, "Abrir Arquivo", "", filters=filters, options=options)
        else:  # PyQt4 or older
            file_name, _ = QFileDialog.getOpenFileName(self, "Abrir Arquivo", "", "CSV Files (*.csv);;Excel Files (*.xlsx)", options=options)

        if file_name:
            try:
                if file_name.endswith('.csv'):
                    data = pd.read_csv(file_name)
                elif file_name.endswith(('.xlsx', '.xlsm')):
                    data = pd.read_excel(file_name)
                else:
                    print(f"Formato não suportado: {file_name}")
                    return

                self.table.setColumnCount(len(data.columns))
                self.table.setRowCount(len(data.index))
                self.table.setHorizontalHeaderLabels(data.columns)
                # Preencher a tabela
                for i, row in data.iterrows():
                    for j, value in enumerate(row):
                        if isinstance(value, (int, float)):
                            item = QTableWidgetItem(f"{value:.2f}" if pd.notnull(value) else " ") # Formatar com duas casas decimais

                        else:
                            item = QTableWidgetItem(str(value) if pd.notnull(value) else " ")
                        self.table.setItem(i, j, item)

                # Extrair valores únicos da coluna E
                if len(data.columns) >= 5:
                    bairros = data.iloc[:, 8].astype(str).unique()  # Converter para string e obter valores únicos
                    bairros = [b for b in bairros if not b.isnumeric()]  # Filtrar valores numéricos
                    self.combobox.addItems(bairros)
                    self.combobox.setCurrentIndex(0)

            except Exception as e:
                print(f"Erro ao abrir o arquivo: {e}")
        else:
            print("Nenhum arquivo selecionado.")



    def filtrar(self):
        texto_filtro  = self.combobox.currentText()

        for row in range(self.table.rowCount()):
            item = self.table.item(row, 8)
            if item is not None and texto_filtro.lower() in item.text().lower():
                self.table.showRow(row)
            else:
                self.table.hideRow(row)


    def salvar_dados(self):
        file_name_base = self.line_edit.text()

        workbook = Workbook()
        sheet = workbook.active

        num_rows = self.table.rowCount()
        num_cols = self.table.columnCount()

        for row in range(num_rows):
            if not self.table.isRowHidden(row):  # Verifica se a linha está visível
                for col in range(num_cols):
                    item = self.table.item(row, col)
                    if item is not None:
                        sheet.cell(row=row+1, column=col+1, value=item.text())
                    else:
                        sheet.cell(row=row+1, column=col+1, value="")
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog

        if not file_name_base:
            file_name, _ = QFileDialog.getSaveFileName(self, "Salvar como ...", "", "Excel Files (*.xlsx)", options=options)
        else:
            if not file_name_base.endswith(".xlsx"):
                file_name_base += ".xlsx"
            file_name, _ = QFileDialog.getSaveFileName(self, "Salvar como ...", file_name_base, "Excel Files (*.xlsx)", options=options)

        if file_name:
            workbook.save(filename=file_name)
if __name__ == '__main__':

    
    
    app = QApplication(sys.argv)
    window = MainWindow()
    window.setWindowTitle("Gerenciamento")
    sys.exit(app.exec_())
